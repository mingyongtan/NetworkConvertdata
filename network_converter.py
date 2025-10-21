#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
network_parse.py — TXT ➜ XLSX network stats converter (GUI + CLI)

New in this revision
--------------------
- **Correct CSV handling**: Wireshark exports are comma-separated with quotes.
  The parser now uses Python's `csv` to read commas/tabs/semicolons correctly
  (no more whole-row-in-column A issues). Headers are preserved exactly as in
  the file (e.g., `AS Number`, `AS Organization`).
- Header detection no longer guesses: we trust the first parsed row as header
  after skipping an optional protocol label line (`Ethernet`, `IPv4`, `IPv6`,
  `TCP`, `UDP`).
- Numeric coercion by column name (Packets/Bytes/Port/Latitude/Longitude),
  but without rewriting header titles.
- Empty-file UX unchanged; GUI auto-opens on empty in CLI (`--on-empty gui`).
- **More tests**: added CSV-based Wireshark samples (with quotes + geo/ASN
  columns) to guarantee proper columnization.
- **IPv4/IPv6 geo drop**: For IPv4 and IPv6 sheets, the following columns are
  now removed automatically on export: `Country`, `City`, `Latitude`,
  `Longitude`, `AS Number`, `AS Organization`. Use of GUI/CLI is unchanged.
- **TCP/UDP port drop**: For TCP and UDP sheets, the `Port` column is removed
  automatically during conversion.
- **Pareto**: Rows sorted by per-row %; cumulative closest to 80 within 78–90
  is written to `Top 20 %` (2 decimals) and the cutoff row is highlighted.
- **Multi-folder GUI**: Add multiple folders; each becomes its own workbook.

Supported inputs: Ethernet / IPv4 / IPv6 / TCP / UDP dumps with headers.
Writes an Excel workbook with one sheet per protocol (real Excel Tables).

Usage
-----
# GUI (click-pick files or folders)
python network_parse.py --gui

# CLI — convert current folder
python network_parse.py

# CLI — specific folder or files (globs OK)
python network_parse.py /path/to/folder -o /path/to/out.xlsx
python network_parse.py data/*.txt -o network.xlsx

# Preview only / self-test
python network_parse.py --list-only data/
python network_parse.py --selftest

Dependencies: pandas, openpyxl
   pip install pandas openpyxl
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import glob
import csv
import tempfile
import shutil
from typing import Dict, List, Optional, Tuple

try:
    import pandas as pd
except ImportError as e:
    raise SystemExit("This tool needs pandas. Install with: pip install pandas openpyxl") from e

# -------------------------
# Helpers / constants
# -------------------------
PROTO_NAMES = {"ethernet", "ipv4", "ipv6", "tcp", "udp"}
GEO_DROP_COLS = [
    "Country", "City", "Latitude", "Longitude", "AS Number", "AS Organization",
]

def NORMALIZE(s: Optional[str]) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").strip().lower())

# -------------------------
# Parsing
# -------------------------

def _maybe_drop_proto_label(lines: List[str]) -> List[str]:
    clean = [ln.rstrip("\r\n") for ln in lines]
    # drop leading empties
    while clean and not clean[0].strip():
        clean.pop(0)
    if clean and NORMALIZE(clean[0]) in PROTO_NAMES and "," not in clean[0] and "\t" not in clean[0]:
        return clean[1:]
    return clean


def _csv_parse(lines: List[str]) -> Tuple[List[str], List[List[str]], str]:
    """Return (header, rows, delimiter). Prefers csv.Sniffer, falls back."""
    data = _maybe_drop_proto_label(lines)
    # Remove blank lines; keep content otherwise
    nonempty = [ln for ln in data if ln.strip()]
    if not nonempty:
        return [], [], ","

    sample = "\n".join(nonempty[: min(50, len(nonempty))])
    try:
        sniff = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delim = sniff.delimiter
    except Exception:
        # heuristic: if comma appears a lot, choose comma; else tab; else spaces
        counts = {d: sample.count(d) for d in [",", "\t", ";", "|"]}
        delim = max(counts, key=counts.get) if max(counts.values()) > 0 else ","

    reader = csv.reader(nonempty, delimiter=delim)
    try:
        header = next(reader)
    except StopIteration:
        return [], [], delim

    # strip BOM and whitespace
    if header:
        header[0] = header[0].lstrip("\ufeff").strip()
        header = [h.strip() for h in header]

    rows: List[List[str]] = []
    for row in reader:
        # pad/truncate to header len
        if len(row) < len(header):
            row = row + ["" for _ in range(len(header) - len(row))]
        elif len(row) > len(header):
            row = row[: len(header)]
        rows.append([c.strip() for c in row])

    return header, rows, delim


def detect_header_and_rows(lines: List[str]) -> Tuple[Optional[List[str]], List[List[str]]]:
    """Parse with CSV first; if that yields a header, return it; otherwise fall back."""
    header, rows, _ = _csv_parse(lines)
    if header:
        return header, rows

    # Fallback to old splitter (tabs/2+ spaces)
    clean = [ln.strip("\r\n") for ln in lines if ln.strip()]
    if not clean:
        return None, []

    if NORMALIZE(clean[0]) in PROTO_NAMES:
        clean = clean[1:]

    parts0 = re.split(r"\s{2,}|\t", clean[0].strip())
    if len(parts0) >= 3:  # looks like header-ish
        header = [p.strip() for p in parts0]
        rows = [re.split(r"\s{2,}|\t", ln.strip()) for ln in clean[1:]]
        return header, rows

    # Last resort: treat all lines as data (single column)
    return None, [[x] for x in clean]


def to_dataframe(header: Optional[List[str]], rows: List[List[str]], expected_cols: Optional[List[str]] = None) -> pd.DataFrame:
    """Build a DataFrame. If a header is provided, we **use it as-is**.
    If no header, use expected_cols (if provided) or infer width from data.
    """
    if header:
        df = pd.DataFrame(rows, columns=header)
    else:
        if not rows:
            return pd.DataFrame(columns=(expected_cols or []))
        width = max((len(r) for r in rows), default=1)
        cols = (expected_cols[:width] if expected_cols else [f"Col{i+1}" for i in range(width)])
        df = pd.DataFrame(rows, columns=cols)

    # Coerce numeric types by column name
    NUMERIC_COLS = {
        "packets", "bytes", "txpackets", "txbytes", "rxpackets", "rxbytes",
        "port", "latitude", "longitude", "asnumber",
    }
    for c in df.columns:
        if NORMALIZE(c) in NUMERIC_COLS:
            df[c] = pd.to_numeric(df[c].replace({"": pd.NA}), errors="coerce")

    return df


def parse_file(path: str) -> Tuple[str, pd.DataFrame]:
    """Parse a single .txt/.csv file and return (sheet_name, DataFrame)."""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    # Pick sheet name from filename (fallback to generic)
    basename = os.path.basename(path).lower()
    DISPLAY_NAME = {"ethernet": "Ethernet", "ipv4": "IPv4", "ipv6": "IPv6", "tcp": "TCP", "udp": "UDP"}
    proto = None
    for key, disp in DISPLAY_NAME.items():
        if key in basename:
            proto = disp
            break
    if proto is None:
        proto = "Sheet"

    header, rows = detect_header_and_rows(lines)
    df = to_dataframe(header, rows, expected_cols=None)

    # For IPv4/IPv6, drop Geo/ASN columns if present
    if proto in ("IPv4", "IPv6"):
        drop_cols = [c for c in GEO_DROP_COLS if c in df.columns]
        if drop_cols:
            df = df[[c for c in df.columns if c not in drop_cols]]

    # For TCP/UDP, drop Port column if present (case-insensitive guard)
    if proto in ("TCP", "UDP"):
        port_cols = [c for c in df.columns if NORMALIZE(c) == "port"]
        if port_cols:
            df = df[[c for c in df.columns if c not in port_cols]]

    return proto, df


def parse_folder(folder: str, recursive: bool = False, pattern: str = "*.txt") -> List[Tuple[str, pd.DataFrame]]:
    search = os.path.join(folder, "**", pattern) if recursive else os.path.join(folder, pattern)
    files = sorted(glob.glob(search, recursive=recursive))
    if not files:
        return []
    sheets: List[Tuple[str, pd.DataFrame]] = []
    for p in files:
        if not os.path.isfile(p):
            continue
        sheet, df = parse_file(p)
        if not df.empty:
            sheets.append((sheet, df))
    return sheets

# -------------------------
# Excel writing
# -------------------------

def write_sheets_to_excel(output_path: str, sheets: List[Tuple[str, pd.DataFrame]]) -> None:
    from openpyxl.utils import get_column_letter as _gcl
    from openpyxl.worksheet.table import Table as _Table, TableStyleInfo as _TSI
    from openpyxl import load_workbook as _lb

    # --- Preprocess: sort by per-row % of Packets and compute Pareto sum (closest to 80 within 78..90)
    prepped: List[Tuple[str, pd.DataFrame, Optional[float], Optional[int]]] = []
    for sheet_name, df in sheets:
        # find Packets column
        pkt_col = None
        for c in df.columns:
            if NORMALIZE(str(c)) in ("packets", "packet"):
                pkt_col = c; break
        if pkt_col is None:
            prepped.append((sheet_name, df, None, None))
            continue
        try:
            s = pd.to_numeric(df[pkt_col], errors="coerce").fillna(0)
        except Exception:
            prepped.append((sheet_name, df, None, None))
            continue
        total = float(s.sum()) if len(s) else 0.0
        if total <= 0:
            prepped.append((sheet_name, df, None, None))
            continue
        pct = (s / total) * 100.0
        # sort rows by pct desc
        df_sorted = df.copy()
        df_sorted["__pct_tmp__"] = pct.values
        df_sorted = df_sorted.sort_values("__pct_tmp__", ascending=False, kind="mergesort").drop(columns=["__pct_tmp__"]).reset_index(drop=True)
        # choose cumulative sum closest to 80 within [78,90]
        pct_sorted = pct.sort_values(ascending=False).reset_index(drop=True)
        cumsum = pct_sorted.cumsum()
        band = cumsum[(cumsum >= 78) & (cumsum <= 90)]
        if not band.empty:
            best_idx = int((band - 80).abs().idxmin())
            best = float(cumsum.iloc[best_idx])
        else:
            best_idx = int((cumsum - 80).abs().idxmin())
            best = float(cumsum.iloc[best_idx])
        prepped.append((sheet_name, df_sorted, round(best, 2), best_idx))

    # First write sheets with pandas
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        used = set()
        for sheet_name, df_sorted, _best, _idx in prepped:
            name = sheet_name
            k = 2
            while name in used:
                name = f"{sheet_name}_{k}"; k += 1
            used.add(name)
            df_sorted.to_excel(writer, index=False, sheet_name=name)

    # Re-open and add the three computed columns + create Table; also write Pareto value
    wb = _lb(output_path)
    for (sheet_name, _df_sorted, best_sum, best_idx) in prepped:
        ws = wb[sheet_name]
        # Identify header row and find the Packets column (accept 'Packet' or 'Packets')
        headers = [cell.value if cell.value is not None else "" for cell in ws[1]]
        norm = [NORMALIZE(str(h)) for h in headers]
        pkt_idx = None
        for i, h in enumerate(norm, start=1):
            if h in ("packets", "packet"):
                pkt_idx = i
                break
        if pkt_idx is None:
            # no packets column? skip augmenting this sheet
            max_row = ws.max_row
            max_col = ws.max_column
            if max_row >= 2 and max_col >= 1:
                ref = f"A1:{_gcl(max_col)}{max_row}"
                disp = re.sub(r"[^A-Za-z0-9_]", "_", f"T_{ws.title}")[:31]
                t = _Table(displayName=disp, ref=ref)
                t.tableStyleInfo = _TSI(name="TableStyleMedium9", showRowStripes=True)
                ws.add_table(t)
                ws.freeze_panes = "A2"
            continue

        max_row = ws.max_row
        max_col = ws.max_column

        # New columns appended to the right
        col_total = max_col + 1
        col_pct   = max_col + 2
        col_top20 = max_col + 3

        ws.cell(row=1, column=col_total, value="Total Packets")
        ws.cell(row=1, column=col_pct,   value="Total Packets in 100% (B/D *100)")
        ws.cell(row=1, column=col_top20, value="Top 20 %")

        pkt_letter = _gcl(pkt_idx)
        total_letter = _gcl(col_total)

        # Build formulas
        total_sum_formula = f"=SUM(${pkt_letter}$2:${pkt_letter}${max_row})"
        for r in range(2, max_row + 1):
            # Total Packets
            ws.cell(row=r, column=col_total).value = total_sum_formula
            # Total Packets in 100% (B/D *100)
            cell_pct = ws.cell(row=r, column=col_pct)
            cell_pct.value = f"=({pkt_letter}{r}/{total_letter}{r})*100"
            cell_pct.number_format = '0.00'
            # Top 20 % left blank per-row; we'll set row 2 only below

        # If we computed a best Pareto sum, write it once at row 2 of Top 20 %
        if best_sum is not None and max_row >= 2:
            cell_top = ws.cell(row=2, column=col_top20)
            cell_top.value = best_sum
            cell_top.number_format = '0.00'

        # Highlight the row that reached the Pareto cutoff (index -> Excel row)
        if best_idx is not None:
            from openpyxl.styles import PatternFill
            row_to_mark = int(best_idx) + 2  # 1 header + 1-based row index
            fill = PatternFill(fill_type="solid", start_color="00FFF2CC", end_color="00FFF2CC")
            for c in range(1, max_col + 3 + 1):  # include newly added 3 cols
                ws.cell(row=row_to_mark, column=c).fill = fill

        # Create/resize the table to include new columns
        new_max_col = ws.max_column
        ref = f"A1:{_gcl(new_max_col)}{max_row}"
        disp = re.sub(r"[^A-Za-z0-9_]", "_", f"T_{ws.title}")[:31]
        if disp in ws.tables:
            del ws.tables[disp]
        t = _Table(displayName=disp, ref=ref)
        t.tableStyleInfo = _TSI(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(t)
        ws.freeze_panes = "A2"

        # crude autofit for all columns (including new ones)
        for i in range(1, new_max_col + 1):
            col_letter = _gcl(i)
            max_len = 0
            for cell in ws[col_letter]:
                val = cell.value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = min(max(10, int(max_len * 1.05)), 60)

    wb.save(output_path)

# -------------------------
# GUI
# -------------------------

def run_gui() -> int:
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk
    except Exception as e:
        print("GUI unavailable:", e)
        return 2

    class App(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("TXT → XLSX (Network Converter)")
            self.geometry("760x520")
            self.minsize(660, 460)
            self.selected_paths: List[str] = []      # files
            self.selected_folders: List[str] = []    # folders (multi workbook mode)
            self._build_ui()

        def _build_ui(self):
            pad = {"padx": 10, "pady": 8}

            frm_top = ttk.Frame(self); frm_top.pack(fill="x", **pad)
            ttk.Label(frm_top, text="1) Pick .txt files or add folders (one workbook per folder)").pack(anchor="w")
            btns = ttk.Frame(frm_top); btns.pack(fill="x", pady=4)
            ttk.Button(btns, text="Select .txt files", command=self.on_pick_files).pack(side="left")
            ttk.Button(btns, text="Select folder", command=self.on_pick_folder).pack(side="left", padx=8)
            ttk.Button(btns, text="Add subfolders…", command=self.on_pick_parent_subfolders).pack(side="left")
            ttk.Button(btns, text="Clear list", command=self.on_clear).pack(side="left")

            self.lst = tk.Listbox(self, height=10); self.lst.pack(fill="both", expand=False, **pad)

            frm_out = ttk.Frame(self); frm_out.pack(fill="x", **pad)
            ttk.Label(frm_out, text="2) Output (.xlsx for single workbook OR choose a folder for multi-folder mode):").pack(anchor="w")
            self.out_entry = ttk.Entry(frm_out)
            default_out = os.path.join(os.path.expanduser("~"), "Desktop", "network_data.xlsx")
            try: self.out_entry.insert(0, default_out)
            except Exception: self.out_entry.insert(0, os.path.abspath("network_data.xlsx"))
            self.out_entry.pack(fill="x")
            ttk.Button(frm_out, text="Browse...", command=self.on_pick_output).pack(anchor="e", pady=6)

            frm_run = ttk.Frame(self); frm_run.pack(fill="x", **pad)
            ttk.Button(frm_run, text="Convert →", command=self.on_convert).pack(side="left")
            ttk.Button(frm_run, text="Quit", command=self.destroy).pack(side="right")

            ttk.Separator(self).pack(fill="x", pady=6)
            ttk.Label(self, text="Log:").pack(anchor="w", padx=10)
            self.log = tk.Text(self, height=12); self.log.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        def logln(self, msg: str):
            self.log.insert("end", msg + "\n"); self.log.see("end"); self.update_idletasks()

        def refresh_listbox(self):
            self.lst.delete(0, "end")
            for d in self.selected_folders:
                self.lst.insert("end", f"[DIR] {d}")
            for p in self.selected_paths:
                self.lst.insert("end", p)

        def on_pick_files(self):
            paths = filedialog.askopenfilenames(title="Select .txt files", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
            if not paths: return
            self.selected_paths.extend(list(paths))
            self.selected_paths = list(dict.fromkeys(self.selected_paths))
            self.refresh_listbox()

        def on_pick_folder(self):
            folder = filedialog.askdirectory(title="Select folder with .txt files")
            if not folder: return
            if folder not in self.selected_folders:
                self.selected_folders.append(folder)
            self.refresh_listbox()

        def on_pick_parent_subfolders(self):
            parent = filedialog.askdirectory(title="Select a parent folder (adds its immediate subfolders)")
            if not parent: return
            try:
                subs = [os.path.join(parent, d) for d in os.listdir(parent) if os.path.isdir(os.path.join(parent, d))]
            except Exception as e:
                messagebox.showerror("Browse failed", str(e)); return
            if not subs:
                messagebox.showinfo("No subfolders", "That folder has no subfolders to add."); return
            added = 0
            for d in subs:
                if d not in self.selected_folders:
                    self.selected_folders.append(d); added += 1
            self.refresh_listbox()
            if added:
                self.logln(f"➕ Added {added} subfolder(s) from: {parent}")
            else:
                self.logln("No new subfolders were added (duplicates skipped).")

        def on_clear(self):
            self.selected_paths = []
            self.selected_folders = []
            self.refresh_listbox()

        def on_pick_output(self):
            # If folders were selected, choose an output directory to drop multiple workbooks
            if self.selected_folders:
                d = filedialog.askdirectory(title="Choose output folder for workbooks")
                if d:
                    self.out_entry.delete(0, "end"); self.out_entry.insert(0, d)
                return
            initial = self.out_entry.get().strip() or "network_data.xlsx"
            path = filedialog.asksaveasfilename(title="Save Excel workbook", defaultextension=".xlsx", initialfile=os.path.basename(initial), filetypes=[("Excel Workbook", "*.xlsx")])
            if path:
                if not path.lower().endswith(".xlsx"): path += ".xlsx"
                self.out_entry.delete(0, "end"); self.out_entry.insert(0, path)

        def on_convert(self):
            folder_mode = bool(self.selected_folders)
            if not self.selected_paths and not folder_mode:
                messagebox.showwarning("No input", "Add .txt files or folders first."); return
            out = self.out_entry.get().strip()

            if folder_mode:
                # Need an output directory
                if not out or not os.path.isdir(out):
                    d = filedialog.askdirectory(title="Choose output folder for workbooks")
                    if not d:
                        messagebox.showwarning("No output folder", "Please choose an output folder for the generated workbooks.")
                        return
                    out = d
                total_workbooks = 0
                for folder in self.selected_folders:
                    try:
                        sheets = parse_folder(folder)
                        if not sheets:
                            self.logln(f"⚠️ {os.path.basename(folder)} → no .txt files found (skipped)")
                            continue
                        name = os.path.basename(os.path.normpath(folder)) or "workbook"
                        out_path = os.path.join(out, f"{name}.xlsx")
                        write_sheets_to_excel(out_path, sheets)
                        self.logln(f"✅ Saved {out_path}")
                        total_workbooks += 1
                    except Exception as e:
                        self.logln(f"❌ {folder} → {e}")
                if total_workbooks:
                    messagebox.showinfo("Done", f"Created {total_workbooks} workbook(s) in: {out}")
                else:
                    messagebox.showwarning("Nothing created", "No workbooks were generated.")
                return

            # Single workbook from selected files (existing behavior)
            if not out:
                messagebox.showwarning("No output", "Enter an output .xlsx path."); return

            sheets: List[Tuple[str, pd.DataFrame]] = []
            ok = 0
            for p in self.selected_paths:
                try:
                    proto, df = parse_file(p)
                    if df.empty:
                        self.logln(f"⚠️ {os.path.basename(p)} → parsed 0 rows (skipped)"); continue
                    sheets.append((proto, df))
                    self.logln(f"✅ {os.path.basename(p)} → '{proto}' with {len(df)} rows"); ok += 1
                except Exception as e:
                    self.logln(f"❌ {os.path.basename(p)} → {e}")

            if not sheets:
                messagebox.showerror("Nothing to write", "No data was parsed from the selected files."); return

            try:
                write_sheets_to_excel(out, sheets)
                self.logln(f"💾 Saved workbook → {out}")
                messagebox.showinfo("Done", f"Converted {ok} file(s).\nSaved to:\n{out}")
            except Exception as e:
                self.logln(f"❌ Failed to save workbook → {e}"); messagebox.showerror("Save failed", str(e))

    if sys.platform.startswith("win"):
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

    app = App(); app.mainloop(); return 0

# -------------------------
# CLI
# -------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Convert network .txt files to an Excel workbook",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python network_parse.py                # current folder -> network_data.xlsx\n"
            "  python network_parse.py --gui         # launch GUI\n"
            "  python network_parse.py data/ -o out.xlsx\n"
            "  python network_parse.py data/*.txt -o out.xlsx\n"
            "  python network_parse.py --list-only data/\n"
            "  python network_parse.py --selftest\n"
        ),
    )

    p.add_argument("inputs", nargs="*", help="Folder path (converts *.txt) OR one or more .txt files (globs OK). Default: .", default=["."])
    p.add_argument("-o", "--output", default="network_data.xlsx", help="Output .xlsx path (default: network_data.xlsx)")
    p.add_argument("-r", "--recursive", action="store_true", help="Recurse into subdirectories when a folder is provided.")
    p.add_argument("-p", "--pattern", default="*.txt", help="Glob pattern to match files inside folders (default: *.txt)")
    p.add_argument("--list-only", action="store_true", help="List files that would be parsed and exit (no Excel writing)")
    p.add_argument("--selftest", action="store_true", help="Run built-in tests (creates temporary input files, checks output)")
    p.add_argument("--on-empty", choices=["gui", "recursive", "error", "selftest", "noop"], default="gui", help=(
        "What to do if no files are found: \n"
        "  gui       = open a file picker (default)\n"
        "  recursive = retry search with --recursive and proceed if found\n"
        "  error     = print a brief message and exit 2\n"
        "  selftest  = run built-in test and exit 0\n"
        "  noop      = print a notice and exit 0\n"
    ))
    p.add_argument("--gui", action="store_true", help="Launch the Tkinter GUI instead of CLI")
    return p


def resolve_inputs(inputs: List[str], recursive: bool, pattern: str) -> List[Tuple[str, pd.DataFrame]]:
    if len(inputs) == 1 and os.path.isdir(inputs[0]):
        return parse_folder(inputs[0], recursive=recursive, pattern=pattern)
    paths: List[str] = []
    for it in inputs:
        if os.path.isdir(it):
            paths.extend(sorted(glob.glob(os.path.join(it, "**", pattern) if recursive else os.path.join(it, pattern), recursive=recursive)))
        else:
            matches = glob.glob(it); paths.extend(matches if matches else [it])
    paths = [p for p in paths if os.path.isfile(p)]
    if not paths: return []
    sheets: List[Tuple[str, pd.DataFrame]] = []
    for p in paths:
        sheet, df = parse_file(p)
        if not df.empty: sheets.append((sheet, df))
    return sheets


def _print_no_files_help(inputs: List[str], pattern: str, recursive: bool) -> None:
    hint_dir = inputs[0] if inputs else "."; rec = " (searched subfolders)" if recursive else ""
    print(f"No matching files found in: {hint_dir}{rec}  |  pattern: {pattern}")

# -------------------------
# Self-test (adds real test cases)
# -------------------------

def _write_sample_tabs(dir_: str) -> None:
    samples = {
        "Ethernet.txt": """Ethernet\nAddress\tPort\tPackets\tBytes\tTx Packets\tTx Bytes\tRx Packets\tRx Bytes\n10.0.0.2\t53\t420\t128000\t210\t64000\t210\t64000\n10.0.0.3\t80\t300\t96000\t150\t48000\t150\t48000\n""",
        "IPv4.txt": """IPv4\nAddress\tPackets\tBytes\tTx Packets\tTx Bytes\tRx Packets\tRx Bytes\n172.16.0.5\t1000\t320000\t600\t200000\t400\t120000\n192.168.1.77\t250\t80000\t150\t48000\t100\t32000\n""",
        "IPv6.txt": """IPv6\nAddress\tPackets\tBytes\tTx Packets\tTx Bytes\tRx Packets\tRx Bytes\n2001:db8::10\t900\t288000\t500\t160000\t400\t128000\nfe80::1\t120\t38000\t60\t19000\t60\t19000\n""",
        "TCP.txt": """TCP\nAddress\tPort\tPackets\tBytes\tTx Packets\tTx Bytes\tRx Packets\tRx Bytes\n203.0.113.9\t443\t2000\t640000\t1200\t384000\t800\t256000\n198.51.100.44\t22\t300\t96000\t160\t51200\t140\t44800\n""",
        "UDP.txt": """UDP\nAddress\tPort\tPackets\tBytes\tTx Packets\tTx Bytes\tRx Packets\tRx Bytes\n8.8.8.8\t53\t800\t256000\t500\t160000\t300\t96000\n224.0.0.251\t5353\t120\t38000\t80\t25000\t40\t13000\n""",
    }
    for name, content in samples.items():
        with open(os.path.join(dir_, name), "w", encoding="utf-8") as f: f.write(content)


def _write_sample_csv(dir_: str) -> None:
    # CSV with quotes and geo/ASN columns (mirrors Wireshark export)
    ipv4_csv = (
        'Address,Packets,Bytes,Tx Packets,Tx Bytes,Rx Packets,Rx Bytes,Country,City,Latitude,Longitude,AS Number,AS Organization\n'
        '"3.169.252.32",1407,1502693,1017,1464783,390,37910,US,"New York",40.7128,-74.0060,15169,"Google LLC"\n'
        '"3.175.96.41",170,90357,90,63582,80,26775,US,"New York",40.7128,-74.0060,14618,"Amazon"\n'
    )
    ipv6_csv = (
        'Address,Packets,Bytes,Tx Packets,Tx Bytes,Rx Packets,Rx Bytes,Country,City,Latitude,Longitude,AS Number,AS Organization\n'
        '"fe80::1",5,450,0,0,5,450,US,"",0,0,0,""\n'
    )
    with open(os.path.join(dir_, "IPv4_csv.txt"), "w", encoding="utf-8") as f: f.write(ipv4_csv)
    with open(os.path.join(dir_, "IPv6_csv.txt"), "w", encoding="utf-8") as f: f.write(ivp6:=ipv6_csv)


def run_selftest() -> None:
    tmpdir = tempfile.mkdtemp(prefix="netparse_")
    try:
        # Test A: tab-separated samples
        _write_sample_tabs(tmpdir)
        out = os.path.join(tmpdir, "test_out.xlsx")
        sheets = parse_folder(tmpdir)
        assert sheets and len(sheets) == 5, "Expected 5 parsed sheets from tab samples"
        # Ensure TCP/UDP had Port dropped
        d = {name: df for name, df in sheets}
        if "TCP" in d:
            assert "Port" not in d["TCP"].columns, "TCP sheet should not contain Port"
        if "UDP" in d:
            assert "Port" not in d["UDP"].columns, "UDP sheet should not contain Port"
        write_sheets_to_excel(out, sheets)
        assert os.path.isfile(out), "Output workbook not created"

        # Test B: explicit file list yields same count
        files = sorted(glob.glob(os.path.join(tmpdir, "*.txt")))
        sheets2 = resolve_inputs(files, recursive=False, pattern="*.txt")
        assert len(sheets2) == len(sheets) == 5, "Expected 5 sheets from samples (file-list path)"

        # Test C: CSV Wireshark-like samples (with quotes + geo columns)
        _write_sample_csv(tmpdir)
        ipv4_path = os.path.join(tmpdir, "IPv4_csv.txt")
        proto, df_ipv4 = parse_file(ipv4_path)
        assert not df_ipv4.empty and proto == "IPv4"
        for col in GEO_DROP_COLS:
            assert col not in df_ipv4.columns, f"{col} should be dropped for IPv4"
        for col in ["Address", "Packets", "Bytes", "Tx Packets", "Tx Bytes", "Rx Packets", "Rx Bytes"]:
            assert col in df_ipv4.columns
        assert pd.api.types.is_numeric_dtype(df_ipv4["Packets"]) and pd.api.types.is_numeric_dtype(df_ipv4["Bytes"])  # coerced numerics

        # IPv6 CSV case also drops geo/ASN
        ipv6_path = os.path.join(tmpdir, "IPv6_csv.txt")
        proto6, df_ipv6 = parse_file(ipv6_path)
        assert not df_ipv6.empty and proto6 == "IPv6"
        for col in GEO_DROP_COLS:
            assert col not in df_ipv6.columns, f"{col} should be dropped for IPv6"

        # Test D: empty directory returns [] for both helpers (no crash)
        empty_dir = tempfile.mkdtemp(prefix="netparse_empty_")
        try:
            sheets_empty = parse_folder(empty_dir); assert sheets_empty == []
            sheets_empty2 = resolve_inputs([empty_dir], recursive=False, pattern="*.txt"); assert sheets_empty2 == []
        finally:
            shutil.rmtree(empty_dir, ignore_errors=True)

        # Test E: multi-folder workflow (simulate GUI: one workbook per subfolder)
        parent = os.path.join(tmpdir, "parent"); os.makedirs(parent, exist_ok=True)
        subA = os.path.join(parent, "A"); subB = os.path.join(parent, "B")
        os.makedirs(subA, exist_ok=True); os.makedirs(subB, exist_ok=True)
        _write_sample_tabs(subA); _write_sample_tabs(subB)
        out_dir = os.path.join(tmpdir, "out"); os.makedirs(out_dir, exist_ok=True)
        for sub in (subA, subB):
            sheets_sub = parse_folder(sub)
            assert sheets_sub and len(sheets_sub) == 5
            out_path = os.path.join(out_dir, os.path.basename(sub) + ".xlsx")
            write_sheets_to_excel(out_path, sheets_sub)
            assert os.path.isfile(out_path)

        print("SELFTEST OK →", out)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

# -------------------------
# Main
# -------------------------

def main(argv: Optional[List[str]] = None) -> int:
    args = build_arg_parser().parse_args(argv)

    if args.selftest:
        run_selftest(); return 0
    if args.gui:
        return run_gui()

    sheets = resolve_inputs(args.inputs, recursive=args.recursive, pattern=args.pattern)

    if not sheets:
        strat = args.on_empty
        if strat == "gui":
            print("[info] No files found via CLI; opening GUI to choose files…"); return run_gui()
        if strat == "recursive" and not args.recursive:
            sheets = resolve_inputs(args.inputs, recursive=True, pattern=args.pattern)
            if not sheets:
                _print_no_files_help(args.inputs, args.pattern, True); return 2
        elif strat == "selftest":
            print("[info] No files found; running built-in selftest instead."); run_selftest(); return 0
        elif strat == "noop":
            print("[info] No files found; nothing to do (noop).\n"); return 0
        else:
            _print_no_files_help(args.inputs, args.pattern, args.recursive); return 2

    if args.list_only:
        count = sum(len(df) for _, df in sheets)
        print("Will build sheets (preview):")
        for name, df in sheets: print(f"  - {name}: {len(df)} rows, {len(df.columns)} cols")
        print(f"Total rows across sheets: {count}"); return 0

    out = args.output
    parent = os.path.dirname(os.path.abspath(out))
    if parent and not os.path.exists(parent): os.makedirs(parent, exist_ok=True)

    write_sheets_to_excel(out, sheets)
    print(f"Saved workbook → {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
